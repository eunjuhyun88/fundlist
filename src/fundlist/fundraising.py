from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sqlite3
import urllib.error
import urllib.request
import zlib
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from .store import ensure_parent_dir


FUNDRAISE_FILE_CANDIDATES = [
    "/Users/ej/Downloads/문서/VC_Fundraising/2025-2026 Fund Raising.xlsx",
    "/Users/ej/Downloads/2025-2026 Fund Raising - Web2 VC.csv",
    "/Users/ej/Downloads/2025-2026 Fund Raising - Web3 VC.csv",
    "/Users/ej/Downloads/2025-2026 Fund Raising - grants program.csv",
    "/Users/ej/Downloads/2025-2026 Fund Raising - Accelerator Program.csv",
    "/Users/ej/Downloads/2025-2026 Fund Raising - contact.csv",
    "/Users/ej/Downloads/2025-2026 Fund Raising - Email From Pitchdeck.tsv",
    "/Users/ej/Downloads/2025-2026 Fund Raising.xlsx",
]


def _existing_default_fundraise_files() -> List[str]:
    existing = [path for path in FUNDRAISE_FILE_CANDIDATES if Path(path).exists()]
    return existing if existing else FUNDRAISE_FILE_CANDIDATES


DEFAULT_FUNDRAISE_FILES = _existing_default_fundraise_files()


EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
URL_RE = re.compile(r"https?://[^\s,)\]]+")
DATE_TEXT_RE = re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b")
PDF_OBJECT_RE = re.compile(rb"(\d+)\s+(\d+)\s+obj(.*?)endobj", re.S)
PDF_FONT_REF_RE = re.compile(rb"/(Font\d+)\s+(\d+)\s+0\s+R")
PDF_TOUNICODE_RE = re.compile(rb"/ToUnicode\s+(\d+)\s+0\s+R")
PDF_STREAM_RE = re.compile(rb"stream\r?\n(.*?)endstream", re.S)
PDF_TOKEN_RE = re.compile(
    r"/(?:[^\s/\[\]()<>]+)|\[(?:.|\n|\r)*?\]|\((?:\\.|[^\\)])*\)|<[^>]*>|-?\d+(?:\.\d+)?|[A-Za-z\*']+|\""
)
PDF_TEXT_ITEM_RE = re.compile(r"\((?:\\.|[^\\)])*\)|<[^>]*>")
PDF_NUMERIC_RE = re.compile(r"^[\s$€£₩¥+-]*[\d,]+(?:\.\d+)?%?$")


@dataclass
class FundraisingRecord:
    source_file: str
    source_row: int
    category: str
    org_name: str
    contact_name: str
    email: str
    website: str
    status: str
    region: str
    funding: str
    date_text: str
    notes: str
    raw_json: str

    @property
    def fingerprint(self) -> str:
        raw = "|".join(
            [
                self.source_file,
                str(self.source_row),
                self.category,
                self.org_name,
                self.contact_name,
                self.email,
                self.website,
                self.status,
                self.region,
                self.funding,
                self.date_text,
                self.notes,
            ]
        )
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize_text(value: str) -> str:
    return re.sub(r"[^0-9a-zA-Z가-힣]+", "", value.strip().lower())


def sanitize(value: str, limit: int = 1000) -> str:
    cleaned = value.strip().replace("\x00", "")
    cleaned = re.sub(r"[\x01-\x08\x0b\x0c\x0e-\x1f\x7f]+", " ", cleaned)
    return cleaned[:limit]


def category_from_path(path: Path) -> str:
    name = path.name.lower()
    if "web2 vc" in name:
        return "web2_vc"
    if "web3 vc" in name:
        return "web3_vc"
    if "grants program" in name:
        return "grants_program"
    if "accelerator program" in name:
        return "accelerator_program"
    if "contact" in name:
        return "vc_contact"
    if "email from pitchdeck" in name:
        return "pitchdeck_email"
    if name.endswith(".xlsx"):
        return "xlsx_sheet"
    if name.endswith(".pdf"):
        if any(token in name for token in ["investment", "portfolio", "vesting", "track_record"]):
            return "pdf_investment_list"
        return "pdf_document"
    return "unknown"


class FundraisingStore:
    def __init__(self, db_path: str) -> None:
        ensure_parent_dir(db_path)
        self.conn = sqlite3.connect(db_path)
        self._init_schema()

    def _init_schema(self) -> None:
        self.conn.execute(
            """
            CREATE TABLE IF NOT EXISTS fundraising_records (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              source_file TEXT NOT NULL,
              source_row INTEGER NOT NULL,
              category TEXT NOT NULL,
              org_name TEXT NOT NULL,
              contact_name TEXT NOT NULL,
              email TEXT NOT NULL,
              website TEXT NOT NULL,
              status TEXT NOT NULL,
              region TEXT NOT NULL,
              funding TEXT NOT NULL,
              date_text TEXT NOT NULL,
              notes TEXT NOT NULL,
              raw_json TEXT NOT NULL,
              fingerprint TEXT NOT NULL UNIQUE,
              imported_at TEXT NOT NULL DEFAULT (datetime('now'))
            )
            """
        )
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_fundraise_cat ON fundraising_records(category, imported_at DESC)"
        )
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_fundraise_email ON fundraising_records(email)"
        )
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_fundraise_org ON fundraising_records(org_name)"
        )
        self.conn.commit()

    def insert_records(self, records: Sequence[FundraisingRecord]) -> int:
        sql = """
            INSERT OR IGNORE INTO fundraising_records (
              source_file, source_row, category, org_name, contact_name, email, website,
              status, region, funding, date_text, notes, raw_json, fingerprint
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        before = self.conn.total_changes
        with self.conn:
            self.conn.executemany(
                sql,
                [
                    (
                        r.source_file,
                        r.source_row,
                        r.category,
                        r.org_name,
                        r.contact_name,
                        r.email,
                        r.website,
                        r.status,
                        r.region,
                        r.funding,
                        r.date_text,
                        r.notes,
                        r.raw_json,
                        r.fingerprint,
                    )
                    for r in records
                ],
            )
        return self.conn.total_changes - before

    def stats(self) -> Dict[str, object]:
        cur = self.conn.cursor()
        total = cur.execute("SELECT COUNT(*) FROM fundraising_records").fetchone()[0]
        uniq_orgs = cur.execute(
            "SELECT COUNT(DISTINCT org_name) FROM fundraising_records WHERE org_name <> ''"
        ).fetchone()[0]
        uniq_emails = cur.execute(
            "SELECT COUNT(DISTINCT email) FROM fundraising_records WHERE email <> ''"
        ).fetchone()[0]
        by_category = cur.execute(
            """
            SELECT category, COUNT(*) AS n
            FROM fundraising_records
            GROUP BY category
            ORDER BY n DESC
            """
        ).fetchall()
        missing_email = cur.execute(
            """
            SELECT category, COUNT(*) AS n
            FROM fundraising_records
            WHERE email = ''
            GROUP BY category
            ORDER BY n DESC
            """
        ).fetchall()
        top_targets = cur.execute(
            """
            SELECT category, org_name, email, website, notes
            FROM fundraising_records
            WHERE (email <> '' OR website <> '') AND org_name <> ''
            ORDER BY category, org_name
            LIMIT 80
            """
        ).fetchall()
        return {
            "total_records": total,
            "unique_orgs": uniq_orgs,
            "unique_emails": uniq_emails,
            "by_category": [{"category": row[0], "count": row[1]} for row in by_category],
            "missing_email_by_category": [{"category": row[0], "count": row[1]} for row in missing_email],
            "top_targets": [
                {
                    "category": row[0],
                    "org_name": row[1],
                    "email": row[2],
                    "website": row[3],
                    "notes": (row[4] or "")[:200],
                }
                for row in top_targets
            ],
        }


def extract_emails(text: str) -> List[str]:
    return sorted(set(email.lower() for email in EMAIL_RE.findall(text or "")))


def extract_urls(text: str) -> List[str]:
    return sorted(set(URL_RE.findall(text or "")))


def detect_header_index(rows: Sequence[List[str]]) -> int:
    hints = [
        "fund",
        "task",
        "program",
        "website",
        "link",
        "date",
        "amount",
        "tier",
        "region",
        "active",
        "notes",
        "현재상태",
        "투자사",
        "이메일",
        "그랜트",
        "조직",
        "펀딩규모",
        "금액",
        "설명",
    ]
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows):
        cleaned = [c.strip() for c in row]
        non_empty = [c for c in cleaned if c]
        if len(non_empty) < 2:
            continue
        row_join = " ".join(cleaned).lower()
        score = sum(1 for h in hints if h in row_join) + len(non_empty) // 4
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def map_row(
    category: str,
    source_file: str,
    source_row: int,
    headers: Sequence[str],
    row: Sequence[str],
) -> Optional[FundraisingRecord]:
    mapping = {normalize_text(h): sanitize(row[i]) if i < len(row) else "" for i, h in enumerate(headers)}
    row_join = " | ".join([sanitize(v, limit=400) for v in row if v.strip()])

    def pick(*aliases: str) -> str:
        for alias in aliases:
            a = normalize_text(alias)
            for k, v in mapping.items():
                if not v:
                    continue
                if k == a or a in k:
                    return v
        return ""

    org_name = pick("fundname", "taskname", "program", "투자사", "조직", "그랜트이름", "name")
    if org_name.isdigit():
        org_name = ""
    contact_name = pick("contactname", "담당자", "이름")
    email_field = pick("email", "이메일", "이메일연락처", "contact", "연락처")
    website = pick("website", "link", "url", "링크")
    status = pick("현재상태", "status", "active", "상태날짜")
    region = pick("region", "location", "분야", "지역")
    funding = pick("amountraisedmusd", "amountraised", "펀딩규모", "금액", "amount")
    date_text = pick("dateannouncedandraised", "date", "상태날짜")
    notes = pick("notes", "설명", "출처", "결과", "제출내용적은링크")

    emails = extract_emails(email_field or row_join)
    urls = extract_urls(website or email_field or row_join)
    email = ",".join(emails[:3]) if emails else ""
    if not website and urls:
        website = urls[0]

    if not any([org_name, email, website, notes, funding, date_text]):
        return None

    return FundraisingRecord(
        source_file=source_file,
        source_row=source_row,
        category=category,
        org_name=sanitize(org_name, limit=240),
        contact_name=sanitize(contact_name, limit=240),
        email=sanitize(email, limit=240),
        website=sanitize(website, limit=500),
        status=sanitize(status, limit=200),
        region=sanitize(region, limit=120),
        funding=sanitize(funding, limit=200),
        date_text=sanitize(date_text, limit=120),
        notes=sanitize(notes or row_join, limit=800),
        raw_json=json.dumps({"headers": list(headers), "row": list(row)}, ensure_ascii=False),
    )


def parse_delimited_file(path: Path) -> List[List[str]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows: List[List[str]] = []
    with path.open("r", newline="", encoding="utf-8-sig") as fp:
        reader = csv.reader(fp, delimiter=delimiter)
        for row in reader:
            rows.append([cell.strip() for cell in row])
    return rows


def parse_pitchdeck_tsv(path: Path) -> List[FundraisingRecord]:
    rows = parse_delimited_file(path)
    out: List[FundraisingRecord] = []
    seen: set[str] = set()
    for idx, row in enumerate(rows, start=1):
        text = " ".join(row)
        for email in extract_emails(text):
            if email in seen:
                continue
            seen.add(email)
            out.append(
                FundraisingRecord(
                    source_file=str(path),
                    source_row=idx,
                    category="pitchdeck_email",
                    org_name="",
                    contact_name="",
                    email=email,
                    website="",
                    status="",
                    region="",
                    funding="",
                    date_text="",
                    notes="email list imported from pitchdeck tsv",
                    raw_json=json.dumps({"row": row}, ensure_ascii=False),
                )
            )
    return out


def parse_csv_like(path: Path) -> List[FundraisingRecord]:
    category = category_from_path(path)
    if category == "pitchdeck_email":
        return parse_pitchdeck_tsv(path)

    rows = parse_delimited_file(path)
    if not rows:
        return []

    header_idx = detect_header_index(rows)
    headers = rows[header_idx] if header_idx < len(rows) else []
    headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]

    out: List[FundraisingRecord] = []
    for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not any(cell.strip() for cell in row):
            continue
        rec = map_row(
            category=category,
            source_file=str(path),
            source_row=i,
            headers=headers,
            row=row,
        )
        if rec:
            out.append(rec)
    return out


def parse_xlsx(path: Path) -> List[FundraisingRecord]:
    try:
        import openpyxl  # type: ignore
    except Exception:  # noqa: BLE001
        print(f"[warn] openpyxl is not installed; skipping xlsx file: {path}")
        return []

    out: List[FundraisingRecord] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell).strip() if cell is not None else "" for cell in row])
        if not rows:
            continue
        header_idx = detect_header_index(rows)
        headers = [h if h else f"col_{i+1}" for i, h in enumerate(rows[header_idx])]
        for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            if not any(cell.strip() for cell in row):
                continue
            rec = map_row(
                category=f"xlsx:{sheet}",
                source_file=str(path),
                source_row=i,
                headers=headers,
                row=row,
            )
            if rec:
                out.append(rec)
    return out


def _extract_pdf_object_stream(body: bytes) -> bytes:
    match = PDF_STREAM_RE.search(body)
    if not match:
        return b""
    data = match.group(1).rstrip(b"\r\n")
    if b"/FlateDecode" in body:
        try:
            data = zlib.decompress(data)
        except Exception:  # noqa: BLE001
            return b""
    return data


def _decode_pdf_hex_text(hex_text: str) -> str:
    try:
        return bytes.fromhex(hex_text).decode("utf-16-be", errors="ignore")
    except Exception:  # noqa: BLE001
        return ""


def _parse_pdf_cmap(text: str) -> Dict[int, str]:
    cmap: Dict[int, str] = {}
    lines = [line.strip() for line in text.splitlines()]
    idx = 0
    while idx < len(lines):
        line = lines[idx]
        match = re.match(r"(\d+)\s+beginbfchar", line)
        if match:
            count = int(match.group(1))
            for row in lines[idx + 1 : idx + 1 + count]:
                item = re.match(r"<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>", row)
                if not item:
                    continue
                cmap[int(item.group(1), 16)] = _decode_pdf_hex_text(item.group(2))
            idx += count + 1
            continue

        match = re.match(r"(\d+)\s+beginbfrange", line)
        if match:
            count = int(match.group(1))
            for row in lines[idx + 1 : idx + 1 + count]:
                item = re.match(r"<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>", row)
                if item:
                    start = int(item.group(1), 16)
                    end = int(item.group(2), 16)
                    dest = int(item.group(3), 16)
                    width = len(item.group(3))
                    for offset, code in enumerate(range(start, end + 1)):
                        cmap[code] = _decode_pdf_hex_text(f"{dest + offset:0{width}X}")
                    continue

                item = re.match(r"<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>\s+\[(.*)\]", row)
                if not item:
                    continue
                start = int(item.group(1), 16)
                targets = re.findall(r"<([0-9A-Fa-f]+)>", item.group(3))
                for offset, target in enumerate(targets):
                    cmap[start + offset] = _decode_pdf_hex_text(target)
            idx += count + 1
            continue
        idx += 1
    return cmap


def _parse_pdf_font_maps(raw: bytes) -> Dict[str, Dict[int, str]]:
    objects = {int(match.group(1)): match.group(3) for match in PDF_OBJECT_RE.finditer(raw)}
    font_obj_by_name: Dict[str, int] = {}
    for body in objects.values():
        if b"/Font" not in body:
            continue
        for font_name, obj_ref in PDF_FONT_REF_RE.findall(body):
            font_obj_by_name[font_name.decode("latin1")] = int(obj_ref)

    out: Dict[str, Dict[int, str]] = {}
    for font_name, obj_ref in font_obj_by_name.items():
        body = objects.get(obj_ref, b"")
        match = PDF_TOUNICODE_RE.search(body)
        if not match:
            continue
        cmap_body = objects.get(int(match.group(1)), b"")
        stream = _extract_pdf_object_stream(cmap_body)
        if not stream:
            continue
        out[font_name] = _parse_pdf_cmap(stream.decode("latin1", errors="ignore"))
    return out


def _pdf_unescape_literal_bytes(value: bytes) -> bytes:
    out = bytearray()
    idx = 0
    while idx < len(value):
        byte = value[idx]
        if byte != 92:
            out.append(byte)
            idx += 1
            continue

        idx += 1
        if idx >= len(value):
            break

        escaped = value[idx]
        escapes = {
            ord("n"): 10,
            ord("r"): 13,
            ord("t"): 9,
            ord("b"): 8,
            ord("f"): 12,
            ord("("): 40,
            ord(")"): 41,
            ord("\\"): 92,
        }
        if escaped in escapes:
            out.append(escapes[escaped])
            idx += 1
            continue

        if 48 <= escaped <= 55:
            octal = bytes([escaped])
            idx += 1
            for _ in range(2):
                if idx < len(value) and 48 <= value[idx] <= 55:
                    octal += bytes([value[idx]])
                    idx += 1
                else:
                    break
            out.append(int(octal, 8))
            continue

        if escaped in (10, 13):
            if escaped == 13 and idx + 1 < len(value) and value[idx + 1] == 10:
                idx += 2
            else:
                idx += 1
            continue

        out.append(escaped)
        idx += 1
    return bytes(out)


def _decode_pdf_text_bytes(value: bytes, cmap: Dict[int, str]) -> str:
    if cmap and len(value) % 2 == 0:
        out: List[str] = []
        for idx in range(0, len(value), 2):
            code = int.from_bytes(value[idx : idx + 2], "big")
            out.append(cmap.get(code, ""))
        joined = "".join(out).replace("\x00", "")
        if joined.strip():
            return joined

    if len(value) % 2 == 0:
        try:
            decoded = value.decode("utf-16-be", errors="ignore").replace("\x00", "")
            if decoded.strip():
                return decoded
        except Exception:  # noqa: BLE001
            pass
    return value.decode("latin1", errors="ignore").replace("\x00", "")


def _extract_pdf_text_chunks(path: Path) -> List[Tuple[float, float, str]]:
    raw = path.read_bytes()
    font_maps = _parse_pdf_font_maps(raw)
    chunks: List[Tuple[float, float, str]] = []

    for match in PDF_OBJECT_RE.finditer(raw):
        stream = _extract_pdf_object_stream(match.group(3))
        if not stream:
            continue
        content = stream.decode("latin1", errors="ignore")
        if "Tf" not in content or ("Tj" not in content and "TJ" not in content):
            continue

        current_font = ""
        current_x = 0.0
        current_y = 0.0
        stack: List[str] = []
        for token in PDF_TOKEN_RE.findall(content):
            if token in {"Tf", "Tm", "Td", "Tj", "TJ", "T*", "'"}:
                if token == "Tf" and len(stack) >= 2 and stack[-2].startswith("/"):
                    current_font = stack[-2][1:]
                elif token == "Tm" and len(stack) >= 6:
                    try:
                        current_x = float(stack[-2])
                        current_y = float(stack[-1])
                    except Exception:  # noqa: BLE001
                        pass
                elif token == "Td" and len(stack) >= 2:
                    try:
                        current_x += float(stack[-2])
                        current_y += float(stack[-1])
                    except Exception:  # noqa: BLE001
                        pass
                elif token == "Tj" and stack:
                    operand = stack[-1]
                    raw_text = b""
                    if operand.startswith("("):
                        raw_text = _pdf_unescape_literal_bytes(operand[1:-1].encode("latin1", errors="replace"))
                    elif operand.startswith("<"):
                        try:
                            raw_text = bytes.fromhex(operand[1:-1])
                        except Exception:  # noqa: BLE001
                            raw_text = b""
                    text = _decode_pdf_text_bytes(raw_text, font_maps.get(current_font, {}))
                    text = sanitize(" ".join(text.split()), limit=2000)
                    if text:
                        chunks.append((round(current_y, 1), round(current_x, 1), text))
                elif token == "TJ" and stack:
                    parts: List[str] = []
                    for item in PDF_TEXT_ITEM_RE.findall(stack[-1]):
                        raw_text = b""
                        if item.startswith("("):
                            raw_text = _pdf_unescape_literal_bytes(item[1:-1].encode("latin1", errors="replace"))
                        elif item.startswith("<"):
                            try:
                                raw_text = bytes.fromhex(item[1:-1])
                            except Exception:  # noqa: BLE001
                                raw_text = b""
                        parts.append(_decode_pdf_text_bytes(raw_text, font_maps.get(current_font, {})))
                    text = sanitize(" ".join("".join(parts).split()), limit=2000)
                    if text:
                        chunks.append((round(current_y, 1), round(current_x, 1), text))
                stack = []
                continue
            stack.append(token)
    return chunks


def _group_pdf_rows(chunks: Sequence[Tuple[float, float, str]]) -> List[List[Tuple[float, str]]]:
    grouped: List[Tuple[float, List[Tuple[float, str]]]] = []
    for y_pos, x_pos, text in sorted(chunks, key=lambda item: (item[0], item[1], item[2])):
        if grouped and abs(grouped[-1][0] - y_pos) <= 1.0:
            grouped[-1][1].append((x_pos, text))
        else:
            grouped.append((y_pos, [(x_pos, text)]))
    return [sorted(row, key=lambda item: item[0]) for _, row in grouped]


def _looks_like_pdf_header(row: Sequence[Tuple[float, str]]) -> bool:
    joined = normalize_text(" ".join(text for _, text in row))
    first = normalize_text(row[0][1]) if row else ""
    strong_hints = [
        "프로젝트",
        "투자금",
        "단가",
        "코인",
        "물량",
        "일자",
        "비고",
        "나머지",
    ]
    if first in {normalize_text("프로젝트"), normalize_text("프로젝트명")}:
        return True
    score = sum(1 for hint in strong_hints if normalize_text(hint) in joined)
    return score >= 2


def _looks_like_pdf_main_row(row: Sequence[Tuple[float, str]]) -> bool:
    if not row:
        return False
    if row[0][0] > 120:
        return False
    return any(any(ch.isalpha() for ch in text) for _, text in row[:2])


def _extract_pdf_date_text(values: Sequence[str]) -> str:
    for value in values:
        match = DATE_TEXT_RE.search(value)
        if match:
            return match.group(0)
        if value.strip().upper() == "N/A":
            return value.strip()
    return ""


def _pick_pdf_funding(values: Sequence[str]) -> str:
    numeric_cells: List[Tuple[float, str]] = []
    for value in values:
        text = value.strip()
        if not text or not PDF_NUMERIC_RE.match(text) or DATE_TEXT_RE.search(text):
            continue
        try:
            amount = float(text.replace(",", "").replace("%", "").replace("$", ""))
        except Exception:  # noqa: BLE001
            continue
        numeric_cells.append((amount, text))

    for amount, text in numeric_cells:
        if amount >= 1000:
            return text
    return numeric_cells[0][1] if numeric_cells else ""


def parse_pdf(path: Path) -> List[FundraisingRecord]:
    category = category_from_path(path)
    rows = _group_pdf_rows(_extract_pdf_text_chunks(path))
    out: List[FundraisingRecord] = []
    pending_notes: List[str] = []
    header_mode = False

    for row in rows:
        values = [text.strip() for _, text in row if text.strip()]
        if not values:
            continue

        if _looks_like_pdf_header(row):
            pending_notes = []
            header_mode = True
            continue

        if not _looks_like_pdf_main_row(row):
            if header_mode:
                continue
            pending_notes.append(" ".join(values))
            continue

        header_mode = False
        org_name = values[0]
        if not org_name or org_name.isdigit():
            pending_notes = []
            continue

        row_join = " | ".join(values)
        emails = extract_emails(row_join)
        urls = extract_urls(row_join)
        notes_parts = pending_notes + values[1:]
        out.append(
            FundraisingRecord(
                source_file=str(path),
                source_row=len(out) + 1,
                category=category,
                org_name=sanitize(org_name, limit=240),
                contact_name="",
                email=sanitize(",".join(emails[:3]), limit=240),
                website=sanitize(urls[0] if urls else "", limit=500),
                status="",
                region="",
                funding=sanitize(_pick_pdf_funding(values[1:]), limit=200),
                date_text=sanitize(_extract_pdf_date_text(values[1:]), limit=120),
                notes=sanitize(" | ".join([part for part in notes_parts if part]), limit=800),
                raw_json=json.dumps(
                    {
                        "pdf_row": [{"x": x_pos, "text": text} for x_pos, text in row],
                        "pending_prefix_notes": pending_notes,
                    },
                    ensure_ascii=False,
                ),
            )
        )
        pending_notes = []

    if pending_notes and out:
        last = out[-1]
        last.notes = sanitize(f"{last.notes} | {' '.join(pending_notes)}", limit=800)
    return out


def import_fundraising_files(db_path: str, files: Sequence[str]) -> Tuple[int, int]:
    store = FundraisingStore(db_path)
    all_records: List[FundraisingRecord] = []

    for raw in files:
        p = Path(raw).expanduser()
        if not p.exists():
            print(f"[warn] file not found: {p}")
            continue
        try:
            if p.suffix.lower() in {".csv", ".tsv"}:
                parsed = parse_csv_like(p)
            elif p.suffix.lower() == ".xlsx":
                parsed = parse_xlsx(p)
            elif p.suffix.lower() == ".pdf":
                parsed = parse_pdf(p)
            else:
                print(f"[warn] unsupported extension skipped: {p}")
                continue
        except Exception as exc:  # noqa: BLE001
            print(f"[warn] parse failed: {p} -> {exc}")
            continue

        all_records.extend(parsed)
        print(f"[ok] imported from {p.name}: {len(parsed)}")

    inserted = store.insert_records(all_records)
    return len(all_records), inserted


def get_hf_token() -> str:
    return (
        os.environ.get("HF_TOKEN", "").strip()
        or os.environ.get("HUGGINGFACE_API_KEY", "").strip()
        or os.environ.get("HUGGINGFACEHUB_API_TOKEN", "").strip()
    )


def get_openrouter_key() -> str:
    return os.environ.get("OPENROUTER_API_KEY", "").strip()


def call_groq_summary(stats: Dict[str, object], model: str) -> str:
    api_key = os.environ.get("GROQ_API_KEY", "").strip()
    if not api_key:
        return "GROQ_API_KEY is not set. AI summary skipped."

    endpoint = os.environ.get("GROQ_BASE_URL", "https://api.groq.com/openai/v1/chat/completions")
    prompt = (
        "You are a fundraising operations analyst. "
        "Write a concise Korean report with: (1) 핵심 인사이트 5개, "
        "(2) 2주 실행 계획, (3) 우선 연락 대상 10개. "
        "Use the provided JSON stats only.\n\n"
        + json.dumps(stats, ensure_ascii=False)
    )
    payload = {
        "model": model,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": "Return practical, execution-ready Korean guidance."},
            {"role": "user", "content": prompt},
        ],
    }

    req = urllib.request.Request(
        endpoint,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8")
    except Exception as exc:  # noqa: BLE001
        return f"AI summary failed: {exc}"

    try:
        parsed = json.loads(body)
        return (
            parsed.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "AI returned empty response.")
        )
    except Exception as exc:  # noqa: BLE001
        return f"AI summary parse failed: {exc}"


def call_gemini_summary(stats: Dict[str, object], model: str) -> str:
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        return "GEMINI_API_KEY is not set. AI summary skipped."

    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    prompt = (
        "You are a fundraising operations analyst. "
        "Write a concise Korean report with: (1) 핵심 인사이트 5개, "
        "(2) 2주 실행 계획, (3) 우선 연락 대상 10개. "
        "Use the provided JSON stats only.\n\n"
        + json.dumps(stats, ensure_ascii=False)
    )
    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": prompt}],
            }
        ],
        "generationConfig": {"temperature": 0.2},
    }

    req = urllib.request.Request(
        endpoint,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8")
    except Exception as exc:  # noqa: BLE001
        return f"AI summary failed: {exc}"

    try:
        parsed = json.loads(body)
        candidates = parsed.get("candidates", [])
        if not candidates:
            return f"AI summary failed: {parsed}"
        content = candidates[0].get("content", {})
        parts = content.get("parts", []) if isinstance(content, dict) else []
        if parts and isinstance(parts[0], dict) and parts[0].get("text"):
            return str(parts[0]["text"])
        return "AI returned empty response."
    except Exception as exc:  # noqa: BLE001
        return f"AI summary parse failed: {exc}"


def call_huggingface_summary(stats: Dict[str, object], model: str) -> str:
    api_key = get_hf_token()
    if not api_key:
        return "HF_TOKEN (or HUGGINGFACE_API_KEY) is not set. AI summary skipped."

    endpoint = os.environ.get("HUGGINGFACE_BASE_URL", "https://router.huggingface.co/v1/chat/completions")
    prompt = (
        "You are a fundraising operations analyst. "
        "Write a concise Korean report with: (1) 핵심 인사이트 5개, "
        "(2) 2주 실행 계획, (3) 우선 연락 대상 10개. "
        "Use the provided JSON stats only.\n\n"
        + json.dumps(stats, ensure_ascii=False)
    )
    router_payload = {
        "model": model,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": "Return practical, execution-ready Korean guidance."},
            {"role": "user", "content": prompt},
        ],
    }

    try:
        req = urllib.request.Request(
            endpoint,
            data=json.dumps(router_payload).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = ""
        try:
            raw = exc.read().decode("utf-8", errors="ignore")
            detail = re.sub(r"<[^>]+>", " ", raw)
            detail = " ".join(detail.split())[:220]
        except Exception:  # noqa: BLE001
            pass
        return f"AI summary failed: HTTP {exc.code}{' - ' + detail if detail else ''}"
    except Exception as exc:  # noqa: BLE001
        return f"AI summary failed: {exc}"

    try:
        parsed = json.loads(body)
        if isinstance(parsed, dict):
            choices = parsed.get("choices")
            if isinstance(choices, list) and choices:
                text = (
                    choices[0].get("message", {}).get("content")
                    if isinstance(choices[0], dict)
                    else None
                )
                if text:
                    return str(text)
            if parsed.get("generated_text"):
                return str(parsed["generated_text"])
            if parsed.get("error"):
                return f"AI summary failed: {parsed.get('error')}"
        if isinstance(parsed, list) and parsed and isinstance(parsed[0], dict):
            if parsed[0].get("generated_text"):
                return str(parsed[0]["generated_text"])
        return "AI returned empty response."
    except Exception as exc:  # noqa: BLE001
        return f"AI summary parse failed: {exc}"


def call_openrouter_summary(stats: Dict[str, object], model: str) -> str:
    api_key = get_openrouter_key()
    if not api_key:
        return "OPENROUTER_API_KEY is not set. AI summary skipped."

    endpoint = os.environ.get("OPENROUTER_BASE_URL", "https://openrouter.ai/api/v1/chat/completions")
    prompt = (
        "You are a fundraising operations analyst. "
        "Write a concise Korean report with: (1) 핵심 인사이트 5개, "
        "(2) 2주 실행 계획, (3) 우선 연락 대상 10개. "
        "Use the provided JSON stats only.\n\n"
        + json.dumps(stats, ensure_ascii=False)
    )
    payload = {
        "model": model,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": "Return practical, execution-ready Korean guidance."},
            {"role": "user", "content": prompt},
        ],
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "https://local.fundlist"),
        "X-Title": os.environ.get("OPENROUTER_APP_TITLE", "fundlist"),
    }

    try:
        req = urllib.request.Request(
            endpoint,
            data=json.dumps(payload).encode("utf-8"),
            headers=headers,
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = ""
        try:
            raw = exc.read().decode("utf-8", errors="ignore")
            detail = re.sub(r"<[^>]+>", " ", raw)
            detail = " ".join(detail.split())[:220]
        except Exception:  # noqa: BLE001
            pass
        return f"AI summary failed: HTTP {exc.code}{' - ' + detail if detail else ''}"
    except Exception as exc:  # noqa: BLE001
        return f"AI summary failed: {exc}"

    try:
        parsed = json.loads(body)
        if isinstance(parsed, dict):
            choices = parsed.get("choices")
            if isinstance(choices, list) and choices and isinstance(choices[0], dict):
                text = choices[0].get("message", {}).get("content")
                if text:
                    return str(text)
            if parsed.get("error"):
                return f"AI summary failed: {parsed.get('error')}"
        return "AI returned empty response."
    except Exception as exc:  # noqa: BLE001
        return f"AI summary parse failed: {exc}"


def write_markdown_report(
    db_path: str,
    output_path: str,
    with_ai: bool,
    model: str,
    ai_provider: str,
) -> str:
    store = FundraisingStore(db_path)
    stats = store.stats()
    out = Path(output_path).expanduser()
    ensure_parent_dir(str(out))

    lines = [
        "# Fundraising Intelligence Report",
        "",
        f"- Generated (UTC): {now_utc_iso()}",
        f"- DB: {db_path}",
        "",
        "## Summary",
        f"- Total records: {stats['total_records']}",
        f"- Unique orgs: {stats['unique_orgs']}",
        f"- Unique emails: {stats['unique_emails']}",
        "",
        "## Category Breakdown",
    ]
    for row in stats["by_category"]:
        lines.append(f"- {row['category']}: {row['count']}")

    lines.extend(["", "## Missing Email By Category"])
    missing = stats["missing_email_by_category"]
    if missing:
        for row in missing:
            lines.append(f"- {row['category']}: {row['count']}")
    else:
        lines.append("- none")

    lines.extend(["", "## Top Outreach Targets"])
    targets = stats["top_targets"][:30]
    if targets:
        for t in targets:
            lines.append(
                f"- [{t['category']}] {t['org_name']} | email={t['email'] or '-'} | site={t['website'] or '-'}"
            )
    else:
        lines.append("- no targets")

    if with_ai:
        provider = ai_provider.strip().lower()
        if provider == "gemini":
            ai_text = call_gemini_summary(stats, model=model)
        elif provider == "huggingface":
            ai_text = call_huggingface_summary(stats, model=model)
        elif provider == "openrouter":
            ai_text = call_openrouter_summary(stats, model=model)
        else:
            ai_text = call_groq_summary(stats, model=model)
            provider = "groq"
        lines.extend(["", f"## AI Summary ({provider})", ai_text])

    out.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return str(out)


def parse_files_argument(files_arg: str) -> List[str]:
    if not files_arg.strip():
        return []
    return [item.strip() for item in files_arg.split(",") if item.strip()]


def resolve_ai_model(provider: str, model: str) -> str:
    raw = (model or "").strip()
    if raw:
        return raw
    p = provider.strip().lower()
    if p == "gemini":
        return os.environ.get("GEMINI_MODEL", "gemini-2.0-flash")
    if p == "huggingface":
        return os.environ.get("HUGGINGFACE_MODEL", "Qwen/Qwen2.5-7B-Instruct")
    if p == "openrouter":
        return os.environ.get("OPENROUTER_MODEL", "openai/gpt-4o-mini")
    return os.environ.get("GROQ_MODEL", "llama-3.3-70b-versatile")


def fundraise_import_command(args: argparse.Namespace) -> int:
    files = parse_files_argument(args.files)
    fetched, inserted = import_fundraising_files(args.db, files)
    print(f"[done] fundraising import parsed={fetched} inserted={inserted} db={args.db}")
    return 0


def fundraise_report_command(args: argparse.Namespace) -> int:
    model = resolve_ai_model(args.ai_provider, args.model)
    path = write_markdown_report(
        db_path=args.db,
        output_path=args.output,
        with_ai=args.with_ai,
        model=model,
        ai_provider=args.ai_provider,
    )
    print(f"[done] report written: {path}")
    return 0


def fundraise_run_command(args: argparse.Namespace) -> int:
    files = parse_files_argument(args.files)
    fetched, inserted = import_fundraising_files(args.db, files)
    print(f"[ok] fundraising import parsed={fetched} inserted={inserted}")
    model = resolve_ai_model(args.ai_provider, args.model)
    path = write_markdown_report(
        db_path=args.db,
        output_path=args.output,
        with_ai=args.with_ai,
        model=model,
        ai_provider=args.ai_provider,
    )
    print(f"[done] report written: {path}")
    return 0
